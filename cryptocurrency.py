from bs4 import BeautifulSoup
from urllib.request import urlopen, Request
import openpyxl as xl
from openpyxl.styles import Font
from twilio.rest import Client 
import keys


url = 'https://cryptoslate.com/coins/'

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

req = Request(url, headers=headers)

webpage = urlopen(req).read()

soup = BeautifulSoup(webpage, 'html.parser')

title = soup.title

mycell = "+19256408522"
TWnumber = '+15075168765'
client = Client(keys.accountSID, keys.authToken)

wb = xl.Workbook()
ws = wb.active
ws.title = 'Top 5 Crypto Currencies'

ws['A1'] = 'Rank'
ws['B1'] = 'Name and Symbol'
ws['C1'] = 'Price'
ws['D1'] = '% Change'
ws["E1"] = 'New Price'
header_font = Font(name='Times New Roman',size=18, bold=True)

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 16

for cell in ws[1:1]:
    cell.font = header_font

table_rows = soup.findAll("tr")

for x in range(1,6):
    tr = table_rows[x].findAll("td")
    rank = tr[0].text
    name = tr[1].text
    price = tr[2].text.replace("$","").replace(",","").replace(" ","")
    daychange = tr[4].text.replace("+","").replace("%","").replace("-","")
    textname = tr[1].text.replace(" ","")
    

    ws["A" + str(x+1) ] = rank
    ws["B" + str(x+1) ] = name
    ws["C" + str(x+1) ] = price
    ws["D" + str(x+1) ] = daychange


    percentchange = float(daychange)/100.00


    if "-" in tr[4]:
        new_price = round(float(price) * (1 - percentchange),2)
    else:
        new_price = round(float(price) * (1+ percentchange),2)


    if textname == 'BitcoinBTC' or textname=='EthereumETH':
        value = float(price) - float(new_price)
        print("yes")
        if value > 5:
            txt = client.messages.create(to=mycell, from_=TWnumber, body=f"Price Change: {name} increased to {new_price}")
        elif value < -5:
            txt = client.messages.create(to=mycell, from_=TWnumber, body=f"Price Change: {name} increased to {new_price}")
        else:
            print()


wb.save('Top 5 Crypto Currencies.xlsx')