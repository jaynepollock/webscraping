import openpyxl as x1
from openpyxl.styles import Font

wb = x1.Workbook()

wb.create_sheet(index=1,title='Second Sheet')

write_sheet = wb['Second Sheet']
read_wb = x1.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']


maxC = read_ws.max_column
maxR = read_ws.max_row

